import os
import shutil
import datetime
from openpyxl import Workbook, load_workbook
import openpyxl
import pandas as pd


# VERIFICANDO O DIA E ANO PARA A PASTA
from datetime import datetime

data_e_hora_atuais = datetime.now()
v_mes_ano_atual = data_e_hora_atuais.strftime('%m_%Y')
print('/' * 70)
print('MES E ANO ATUAL === ' + v_mes_ano_atual)
print('\\' * 70)




# ABRE ARQUIVO TXT 
file = open('CopiandoArquivos.txt', 'r') 
content = file.readlines()

path = ("data/database.xlsx")
wb_obj = openpyxl.load_workbook(path)
# LENDO UMA COLUNA DA PRIMEIRA ABA  
# sheet_obj = wb_obj.active
# ABAIXO DEFINE A ABA ===  worksheets[0]  QUE VAI TRABALHAR
sheet_obj = wb_obj.worksheets[0]


m_row = sheet_obj.max_row
# USO range 2 PARA COMEÇAR A LER DA SEGUNDA LINHA, E column = 1 === PRIMEIRA COLUNA DA ABA
for i in range(2, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    
    # print(cell_obj.value)
    var = (cell_obj.value)
    var = '\\'+ var +'\\MO_PROPRIA\\' + v_mes_ano_atual
    print('VVVVAAARRRRRR === ' + var)
    print('V' * 70)
    print('=' * 70)

    bin_dir = content[0].replace('\n', '')
    bin_dir = bin_dir + var

    print('BIN DIR AAAA ===' , bin_dir)
    print("F" * 70)


    
    path = content[1].replace('\n', '')


    print('PATH LIMPO ===' , path)
    print("-" * 70)

    # path = path + '\\ATC_164\\MO_PROPRIA\\01_2024'
    # path = path + '\\'+ var +'\\MO_PROPRIA\\01_2024'

    # path = (cell_obj.value)
    path = path + '\\'+ var 


    print('PATH ALTERADO === 000' , path)
    print("*" * 70)



    #lista os arquivos a serem copiados
    print("9" * 70)
    print('BIN DIR AAAA ===' , (bin_dir))
    print("9" * 70)

    bin_files = os.listdir(bin_dir)
    bin_files = bin_files






    #verifica se existem arquivos a serem copiados
    if len(bin_files) == 0:
        print('Nenhum arquivo encontrado em %s para ser copiado'
       % (bin_dir))
        print("-" * 70)

    elif len(bin_files) > 0:
        print("ARQUIVOS ENCONTRADOS bin_files ===> ", bin_files)
        print("-" * 70)

        for file in bin_files:
            # print(bin_files)
            # shutil.copy(os.path.join(bin_dir, file), os.path.join(path, file))
            shutil.copy(os.path.join(bin_dir, file), os.path.join(path, file))



            print('ARQUIO ===  %s  === COPIADO PARA A PASTA DESTINO %s' % (file, path))
            print("-" * 70)






    

# \\192.168.15.31\e-notedell\YYYYYY\ATC_164\MO_PROPRIA\01_2024





# bin_dir = content[0].replace('\n', '')
# bin_dir = bin_dir+'\\ATC_164\\MO_PROPRIA\\01_2024'
# # bin_dir = bin_dir+'\\ATC_164\\arquivos'.replace('\n', '')



# print('BIN DIR AAAA ===' , bin_dir)
# print("F" * 70)



# path = content[1].replace('\n', '')


# print('PATH LIMPO ===' , path)
# print("-" * 70)

# path = path + '\\ATC_164\\MO_PROPRIA\\01_2024'


# print('PATH ALTERADO ===' , path)
# print("*" * 70)



# #lista os arquivos a serem copiados
# print("9" * 70)
# print('BIN DIR AAAA ===' , (bin_dir))
# print("9" * 70)

# bin_files = os.listdir(bin_dir)




# bin_files = bin_files 

# #verifica se existem arquivos a serem copiados
# if len(bin_files) == 0:
#         print('Nenhum arquivo encontrado em %s para ser copiado'
#        % (bin_dir))
#         print("-" * 70)

# elif len(bin_files) > 0:
#         print("ARQUIVOS ENCONTRADOS bin_files ===> ", bin_files)
#         print("-" * 70)

#         for file in bin_files:
#             # print(bin_files)
#             # shutil.copy(os.path.join(bin_dir, file), os.path.join(path, file))
#             shutil.copy(os.path.join(bin_dir, file), os.path.join(path, file))



#             print('ARQUIO ===  %s  === COPIADO PARA A PASTA DESTINO %s' % (file, path))
#             print("-" * 70)





